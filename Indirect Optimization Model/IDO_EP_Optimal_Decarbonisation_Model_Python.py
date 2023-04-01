'''
Created on 24th May 2022
Last Modified in March 2023

Indirect Optimization Energy Planning Scenario in Pyomo

This script presents the mathematical optimisation formulation concerning 
carbon-constrained energy planning for a specific geographical region or 
district.
 
Power plants are fuelled by either renewable energy sources e.g., solar, 
hydropower, wind, etc. or fossil-based sources e.g., coal, natural gas, oil, nuclear, etc. 

This energy planning is conducted for several periods (user-specified), with 
each period having its respective demand and CO2 emission limit

The satisfaction of the CO2 load limit is via the deployment of compensatory 
energy, CCS/CCSU technology, alternative low-carbon fuels (replacing fossil fuels), 
negative emission technologies (NETs) and carbon trading.

NETs are either made up of energy-producing NETs (EP-NETs) or energy-consuming 
NETs (EC-NETs). Examples of EP-NETs are biomass and biochar, while EC-NETs
are made up of enhanced weathering, direct air capture etc. 

EP-NETs and EC-NETs produce and consume energy during CO2 removal respectively

Alternative solid and gas fuels have low CO2 intensity, with a potential of 
replacing coal and natural gas in fossil-based power plants respectively

Carbon Trading adds a lot of dimensions which enable the formulation to reduce costs and emissions in an energy system 
Concept of carbon is becoming crucial in controlling the emissions

@author: Gul, Purusothmn, Dr Michael Short

'''
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import pandas as pd
import os
from openpyxl import load_workbook

path = os.chdir(r'C:\Users\pc\Desktop\DECO2 v2023\Indirect Optimization Model')

cwd = os.getcwd()

model = pyo.ConcreteModel()

file_name = r'IDO_EP_CT_Base_User_Interface_Pakistan.xlsx'

model.plant = pd.read_excel(file_name, sheet_name = 'PLANT_DATA', index_col = 0, header = 37, nrows = 7).to_dict()
model.EP = pd.read_excel(file_name, sheet_name = 'ENERGY_PLANNING_DATA', index_col = 0, header = 7).to_dict()
model.Emissions_unit_cost = pd.read_excel(file_name, sheet_name = 'CT_DATA', index_col = 0, header = 1, nrows = 7).to_dict()
model.fuel = pd.read_excel(file_name, sheet_name = 'FUEL_COST_DATA', index_col = 0, header = 12).to_dict()
model.REN_CI = pd.read_excel(file_name, sheet_name = 'RENEWABLE_CI_DATA', index_col = 0, header = 9).to_dict()
model.REN_COST = pd.read_excel(file_name, sheet_name = 'RENEWABLE_COST_DATA', index_col = 0, header = 9).to_dict()
model.CPX_1 = pd.read_excel(file_name, sheet_name = 'CAPEX_DATA_1', index_col = 0, header = 20).to_dict()
model.CPX_2 = pd.read_excel(file_name, sheet_name = 'CAPEX_DATA_2', index_col = 0, header = 20).to_dict()
model.SLD_CI = pd.read_excel(file_name, sheet_name = 'ALT_SOLID_CI', index_col = 0, header = 6).to_dict()
model.SLD_COST = pd.read_excel(file_name, sheet_name = 'ALT_SOLID_COST', index_col = 0, header = 6).to_dict()
model.GAS_CI = pd.read_excel(file_name, sheet_name = 'ALT_GAS_CI', index_col = 0, header = 6).to_dict()
model.GAS_COST = pd.read_excel(file_name, sheet_name = 'ALT_GAS_COST', index_col = 0, header = 6).to_dict()
model.CCS_data = pd.read_excel(file_name, sheet_name = 'CCS_DATA', index_col = 0, header = 12).to_dict()
model.NET_CI = pd.read_excel(file_name, sheet_name = 'NET_CI_DATA', index_col = 0, header = 12).to_dict()
model.NET_COST = pd.read_excel(file_name, sheet_name = 'NET_COST_DATA', index_col = 0, header = 12).to_dict()
model.TIME = pd.read_excel(file_name, sheet_name = 'TECH_IMPLEMENTATION_TIME', index_col = 0, header = 19).to_dict()

wb = load_workbook(file_name)

sheet_1 = wb['PLANT_DATA']

#flag specifies if the problem is for maximizing profits, minimizing emissions or minimizing the budget
flag = sheet_1['B32'].value
#CT defines if the carbon trading is turned on or off
CT = sheet_1['G32'].value
#EB defines the emissions available initially to buy from previous period i = 1-1 = 0
EB = sheet_1['G33'].value
#CT_Percentage_Emissions define the percentage of emissions participating in CT
CT_Percentage_Emissions = sheet_1['G34'].value
#Elec_Price_Inflation define the factor in percentage with which electricity price is to be increased as we move to the next period
Elec_Price_Inflation = sheet_1['G35'].value
#Initial_Elec_Price define the initial electricity price for period 1
Initial_Elec_Price = sheet_1['B34'].value
#Max_Elec_Price define the maximum electricity price possible for the final period
Max_Elec_Price = sheet_1['B36'].value
#numperiods define the number of periods for which the optimization problem is solved
numperiods = sheet_1['B33'].value + 1
#Plant_Profit defines if the plant positive profitability is turned on or off for the optimization study
Plant_Profit = sheet_1['B35'].value
#Print_Result indicates if results will be displayed in user interface or not
Print_Result = sheet_1['G36'].value

sheet_2 = wb['CAPEX_DATA_1']

#AFF defines the annualized factor
AFF = sheet_2['B18'].value
#periods allocates the number to each period
periods = list(range(1,numperiods,1))


def multiperiod_energy_planning(model, i):
    model.S = model.plant.keys() 
          
    #LIST OF VARIABLES
    #This variable determines the deployment of energy sources in power plant s for period i
    model.energy = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the CO2 intensity of energy sources in power plant s with CCS technology 1 for period i
    model.CI_RET_1 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0.088)
    
    #This variable determines the CO2 intensity of energy sources in power plant s with CCS technology 2 for period i
    model.CI_RET_2 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0.233)
    
    #Binary variable for power generation by power plant s for period i
    model.A = pyo.Var(i, model.S, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of CCS technology 1 in power plant s for period i
    model.B = pyo.Var(i, model.S, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of CCS technology 2 in power plant s for period i
    model.C = pyo.Var(i, model.S, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of EP-NETs technology 1 for period i
    model.D = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of EP-NETs technology 2 for period i
    model.E = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of EP-NETs technology 3 for period i
    model.F = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of EC-NETs technology 1 for period i 
    model.G = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of EC-NETs technology 2 for period i 
    model.H = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of EC-NETs technology 3 for period i
    model.I = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of solar renewable energy for period i 
    model.J = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of hydropower renewable energy for period i 
    model.K = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of biomass renewable energy for period i 
    model.L = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of biogas renewable energy for period i 
    model.M = pyo.Var(i, domain = pyo.Binary, initialize = 0)
    
    #Binary variable for the deployment of MSW renewable energy for period i 
    model.N = pyo.Var(i, domain = pyo.Binary, initialize = 0)    
    
    #Binary variable for the deployment of alternative solid-based fuel technology 1 in power plant s for period i 
    model.O = pyo.Var(i, model.S, domain = pyo.Binary, initialize = 0)  
    
    #Binary variable for the deployment of alternative solid-based fuel technology 2 in power plant s for period i 
    model.P = pyo.Var(i, model.S, domain = pyo.Binary, initialize = 0)  
    
    #Binary variable for the deployment of alternative gas-based fuel technology 1 in power plant s for period i 
    model.Q = pyo.Var(i, model.S, domain = pyo.Binary, initialize = 0)  
    
    #Binary variable for the deployment of alternative gas-based fuel technology 2 in power plant s for period i 
    model.R = pyo.Var(i, model.S, domain = pyo.Binary, initialize = 0)  
            
    #This variable represents the deployment of CCS technology 1 in power plant s for period i
    model.CCS_1 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0) 
    
    #This variable represents the deployment of CCS technology 2 in power plant s for period i
    model.CCS_2 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0) 
    
    #This variable determines the net energy available from power plant s without CCS deployment for period i
    model.net_energy = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the net energy available from power plant s with the deployment of CCS technology 1 for period i
    model.net_energy_CCS_1 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the net energy available from power plant s with the deployment of CCS technology 2 for period i
    model.net_energy_CCS_2 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable represents the minimum deployment of EP_NETs technology 1 for period i
    model.EP_NET_1 = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable represents the minimum deployment of EP_NETs technology 2 for period i
    model.EP_NET_2 = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable represents the minimum deployment of EP_NETs technology 3 for period i 
    model.EP_NET_3 = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable represents the minimum deployment of EC_NETs technology 1 for period i 
    model.EC_NET_1 = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable represents the minimum deployment of EC_NETs technology 2 for period i
    model.EC_NET_2 = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable represents the minimum deployment of EC_NETs technology 3 for period i
    model.EC_NET_3 = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of solar renewable energy for period i
    model.REN_SOLAR = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of hydro renewable energy for period i
    model.REN_HYDRO = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of biomass renewable energy for period i
    model.REN_BM = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of biogas renewable energy for period i
    model.REN_BG = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of MSW renewable energy for period i
    model.REN_MSW = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of alternative solid-based fuel technology 1 for coal-based plant s for period i
    model.solid_1 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of alternative solid-based fuel technology 2 for coal-based plant s for period i
    model.solid_2 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of alternative gas-based fuel technology 1 for natural gas-based plant s for period i
    model.gas_1 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable determines the minimum deployment of alternative gas-based fuel technology 2 for natural gas-based plant s for period i
    model.gas_2 = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)

    #This variable determines the revised total CO2 emissions for period i
    model.new_emission = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 23.64428683)

    #This variable determines the total energy cost (fuel and annualized capital) of power plant s for period i
    model.energy_cost = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 16.473)
    
    #This variable determines the fuel energy cost of power plant s for period i
    model.fuel_cost = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0.018)
    
    #This variable determines the annualized capital energy cost of power plant s for period i
    model.annualized_capital_cost = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 16.455)

    #This variable determines the total energy planning cost for period i
    model.sum_cost = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 12089.43882)    
    
    #This variable determines the net energy available from power plant s with CCS deployment for period i
    model.net_energy_CCS = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable shows the emissions bought by each powerplant s from government for period i
    model.CT_emissions = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0) 
    
    #This variable forbids the non-energy producing facilities s to take part in CT for period i
    model.emissions_unit_cost_new = pyo.Var(i, model.S, domain = pyo.NonNegativeReals) 
    
    #This variable shows net emissions from a plant s for period i
    model.net_plant_emissions = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable shows the earnings of a power plant by selling stored emissions
    model.CCS_selling_price = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0)
    
    #This variable shows the availability of tradable carbon emissions between plants
    model.ET_plant = pyo.Var(i, model.S, bounds = (-100000, 100000), initialize = 0) 
    
    #This variable shows the carbon emissions trading between plants
    model.ET_plant_final = pyo.Var(i, model.S, bounds = (-100000, 100000), initialize = 0)
    
    #This variable shows the cost paid/earned by each powerplant s in trading emissions among themselves for period i
    model.CT_cost = pyo.Var(i, model.S, bounds = (-1000000, 1000000), initialize = 0) 
     
    #This variable shows the traded carbon emissions in period i from the previous period i-1
    model.ET_plant_new = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 0) 
    
    #This variable determines the total CO2 emission rights left to trade in next period for period i
    model.CT_balance_emissions = pyo.Var(i, initialize = 0)
    
    #This variable determines the total CO2 emission rights left to trade for a plant s in next period i
    model.Emissions_left = pyo.Var(i, model.S, bounds = (-100000, 100000), initialize = 0)
    
    #This variable predicts the electricity prices for each period i
    model.Electricity_price = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 120)
    
    #This variable shows the overall cost of a plant s for period i
    model.Plant_cost = pyo.Var(i, model.S, initialize = 16.473)
    
    #This variable shows the earnings of a power plant by selling electricty
    model.Plant_earning = pyo.Var(i, model.S, domain = pyo.NonNegativeReals, initialize = 99)
    
    #This variable determines the total energy planning cost for period i
    model.total_profit = pyo.Var(i, domain = pyo.NonNegativeReals, initialize = 7413.561183)
    
    
    #OBJECTIVE FUNCTION
    
    #For the minimum budget objective function, the total cost is minimised, subject to the satisfaction of the CO2 emission limit
    #For the minimum emission objective function, the total emission is minimised, subject to the available budgetary constraint
    #For the maximum profit objective function, the total profit or net cost is maximized, subject to the satisfaction of the CO2 emission limit and allocated budget
    if flag == 'min_budget':
        model.obj = pyo.Objective(expr = sum(model.sum_cost[i] for i in periods), sense = pyo.minimize)
    elif flag == 'min_emission':
        model.obj = pyo.Objective(expr = sum(model.new_emission[i] for i in periods), sense = pyo.minimize)
    else:
        model.obj = pyo.Objective(expr = sum(model.total_profit[i] for i in periods), sense = pyo.maximize)

    
    #CONSTRAINTS
    
    
    #Prior to any energy planning, the total power generation from all power plants should satisfy the regional energy demand for period i
    def demand(model, i, s):
        return model.EP['Demand'][i] == sum(model.energy[i,s] for s in model.S)
        
    model.Cons_1 = pyo.Constraint(i, model.S, rule = demand)
    

	#The deployment of energy source in power plant s should at least satisfy the lower bound for period i
    def lower_bound_energy(model, i, s):
        return model.energy[i,s] >= model.plant[s]['LB'] * model.A[i,s]
        
    model.Cons_2 = pyo.Constraint(i, model.S, rule = lower_bound_energy)
    
    
    #The deployment of energy source in power plant s should at most satisfy the upper bound for period i
    def upper_bound_energy(model, i, s):
        return model.energy[i,s] <= model.plant[s]['UB'] * model.A[i,s]
        
    model.Cons_3 = pyo.Constraint(i, model.S, rule = upper_bound_energy)
    

	#There should not be power generation from power plant s before its commissioning period
    def plant_commission(model, i, s):
        if i < model.plant[s]['ON']:
            return model.energy[i,s] == 0
        else:
            return model.energy[i,s] >= 0
            
    model.Cons_4 = pyo.Constraint(i, model.S, rule = plant_commission)
    
    
    #There should be power generation from power plant s after its decommissioning period
    def plant_decommission(model, i, s):
        if i >= model.plant[s]['OFF']:
            return model.energy[i,s] == 0
        else:
            return model.energy[i,s] >= 0
            
    model.Cons_5 = pyo.Constraint(i, model.S, rule = plant_decommission)
    

    #If power plant s is decomissioned in a period, it should remain decommissioned at later periods
    #However, power plant s should be available for power generation till the period before its decommissioning period
    def energy_constraint(model, i, s):
        if i <= model.plant[s]['OFF'] - 2:
            return model.energy[i+1,s] >= model.energy[i,s]
        else:
            return pyo.Constraint.Skip
        
    model.Cons_6 = pyo.Constraint(i, model.S, rule = energy_constraint)
    
    
    #Calculation of carbon intensity of energy sources with CCS technology 1 in power plant s for period i
    def CCS_CI_1(model, i, s):
        return model.plant[s]['CI'] * (1 - model.CCS_data['RR_1'][i]) / (1 - model.CCS_data['X_1'][i]) == model.CI_RET_1[i,s]
    
    model.Cons_7 = pyo.Constraint(i, model.S, rule = CCS_CI_1)
    
    
    #Calculation of carbon intensity of energy sources with CCS technology 2 in power plant s for period i
    def CCS_CI_2(model, i, s):
        return model.plant[s]['CI'] * (1 - model.CCS_data['RR_2'][i]) / (1 - model.CCS_data['X_2'][i]) == model.CI_RET_2[i,s]
    
    model.Cons_8 = pyo.Constraint(i, model.S, rule = CCS_CI_2)
    
    
    #If selected, the deployment of CCS technology 1 in power plant s is limited by the upper bound of the energy output for period i     
    def CCS_limit_1(model, i, s):
        return model.CCS_1[i,s] <= model.plant[s]['UB'] * model.B[i,s]
        
    model.Cons_9 = pyo.Constraint(i, model.S, rule = CCS_limit_1)
	    
    
    #If selected, the deployment of CCS technology 2 in power plant s is limited by the upper bound of the energy output for period i
    def CCS_limit_2(model, i, s):
        return model.CCS_2[i,s] <= model.plant[s]['UB'] * model.C[i,s]
        
    model.Cons_10 = pyo.Constraint(i, model.S, rule = CCS_limit_2)
    
       
    #The total CCS deployment in power plant s should be equal to summation of deployment of individual types of  CCS technology for period i
    def CCS_total(model, i, s):
        return model.CCS_1[i,s] + model.CCS_2[i,s] <= model.energy[i,s]
        
    model.Cons_11 = pyo.Constraint(i, model.S, rule = CCS_total)
    
    
    #The deployment of CCS technology 1 at later periods should at least match the deployment in the previous period
    def CCS_1_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.CCS_1[i+1,s] >= model.CCS_1[i,s]
        
    model.Cons_12 = pyo.Constraint(i, model.S, rule = CCS_1_constraint)
    
    
    #The deployment of CCS technology 2 at later periods should at least match the deployment in the previous period    
    def CCS_2_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.CCS_2[i+1,s] >= model.CCS_2[i,s]
        
    model.Cons_13 = pyo.Constraint(i, model.S, rule = CCS_2_constraint)
    
    
    #Determine the net energy available from power plant s with CCS technology 1 for period i
    def CCS_1_net_energy(model, i, s):
       return model.CCS_1[i,s] * (1 - model.CCS_data['X_1'][i]) == model.net_energy_CCS_1[i,s]
        
    model.Cons_14 = pyo.Constraint(i, model.S, rule = CCS_1_net_energy)

    
    #Determine the net energy available from power plant s with CCS technology 2 for period i
    def CCS_2_net_energy(model, i, s):
        return model.CCS_2[i,s] * (1 - model.CCS_data['X_2'][i]) == model.net_energy_CCS_2[i,s]
        
    model.Cons_15 = pyo.Constraint(i, model.S, rule = CCS_2_net_energy)  
        
        
    #The deployment of alternative solid fuel technology 1 at later periods should at least match the deployment in the previous period
    def alt_solid_1_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.solid_1[i+1,s] >= model.solid_1[i,s]
        
    model.Cons_16 = pyo.Constraint(i, model.S, rule = alt_solid_1_constraint)
    
    
    #The deployment of alternative solid fuel technology 2 at later periods should at least match the deployment in the previous period    
    def alt_solid_2_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.solid_2[i+1,s] >= model.solid_2[i,s]
        
    model.Cons_17 = pyo.Constraint(i, model.S, rule = alt_solid_2_constraint)
    
    
    #The deployment of alternative gas fuel technology 1 at later periods should at least match the deployment in the previous period
    def alt_gas_1_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.gas_1[i+1,s] >= model.gas_1[i,s]
        
    model.Cons_18 = pyo.Constraint(i, model.S, rule = alt_gas_1_constraint)
    
    
    #The deployment of alternative gas fuel technology 2 at later periods should at least match the deployment in the previous period    
    def alt_gas_2_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.gas_2[i+1,s] >= model.gas_2[i,s]
        
    model.Cons_19 = pyo.Constraint(i, model.S, rule = alt_gas_2_constraint)
    
    
    #The total energy contribution must equal the initially determined energy contribution of power plant s for period i
    def fuel_substitution(model, i, s):
        if 'REN' in model.plant[s].values():
            return model.net_energy[i,s] == model.energy[i,s]
        elif 'NUCLEAR' in model.plant[s].values():
            return model.net_energy[i,s] == model.energy[i,s]
        elif 'NG' in model.plant[s].values():
            return model.net_energy[i,s] + model.CCS_1[i,s] + model.CCS_2[i,s] + model.gas_1[i,s] + model.gas_2[i,s] == model.energy[i,s]
        elif 'COAL' in model.plant[s].values():
            return model.net_energy[i,s] + model.CCS_1[i,s] + model.CCS_2[i,s] + model.solid_1[i,s] + model.solid_2[i,s] == model.energy[i,s]
        else:
            return model.net_energy[i,s] + model.CCS_1[i,s] + model.CCS_2[i,s] == model.energy[i,s]
        
    model.Cons_20 = pyo.Constraint(i, model.S, rule = fuel_substitution)  
    
    
    #If power plant s is fuelled by renewable energy sources, CCS technology 1 would not be deployed for period i
    def no_CCS_1(model, i, s):
        if 'REN' in model.plant[s].values():
            return model.CCS_1[i,s] == 0
        elif 'NUCLEAR' in model.plant[s].values():
            return model.CCS_1[i,s] == 0
        else:
            return pyo.Constraint.Skip
    
    model.Cons_21 = pyo.Constraint(i, model.S, rule = no_CCS_1) 
    
    
    #If power plant s is fuelled by renewable energy sources, CCS technology 1 would not be deployed for period i
    def no_CCS_2(model, i, s):
        if 'REN' in model.plant[s].values():
            return model.CCS_2[i,s] == 0
        elif 'NUCLEAR' in model.plant[s].values():
            return model.CCS_2[i,s] == 0
        else:
            return pyo.Constraint.Skip
    
    model.Cons_22 = pyo.Constraint(i, model.S, rule = no_CCS_2)  
    
    
    #If power plant s is fuelled by renewable energy sources, alternative fuels would not be deployed for period i
    def no_AF(model, i, s):
        if 'REN' in model.plant[s].values():
            return model.gas_1[i,s] + model.gas_2[i,s] + model.solid_1[i,s] + model.solid_2[i,s] == 0
        elif 'NUCLEAR' in model.plant[s].values():
            return model.gas_1[i,s] + model.gas_2[i,s] + model.solid_1[i,s] + model.solid_2[i,s] == 0
        else:
            return pyo.Constraint.Skip
    
    model.Cons_23 = pyo.Constraint(i, model.S, rule = no_AF) 
    
    
    #If power plant s is fuelled by renewable energy sources, alternative fuels would not be deployed for period i
    def no_AF_selection(model, i, s):
        if 'REN' in model.plant[s].values():
            return model.O[i,s] + model.P[i,s] + model.Q[i,s] + model.R[i,s] == 0
        elif 'NUCLEAR' in model.plant[s].values():
            return model.O[i,s] + model.P[i,s] + model.Q[i,s] + model.R[i,s] == 0
        else:
            return pyo.Constraint.Skip
    
    model.Cons_24 = pyo.Constraint(i, model.S, rule = no_AF_selection) 
    
    
    #Technology implementation time for CCS technology 1
    def deployment_CCS_1(model, i, s):
        if model.TIME['CCS_1'][i] == 'NO':
            return model.B[i,s] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_25 = pyo.Constraint(i, model.S, rule = deployment_CCS_1)   
    
    
    #Technology implementation time for CCS technology 2
    def deployment_CCS_2(model, i, s):
        if model.TIME['CCS_2'][i] == 'NO':
            return model.C[i,s] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_26 = pyo.Constraint(i, model.S, rule = deployment_CCS_2)   
        
    
    #Technology implementation time for EP-NETs technology 1
    def deployment_EP_NETs_1(model, i):
        if model.TIME['EP_NETs_1'][i] == 'NO':
            return model.D[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_27 = pyo.Constraint(i, rule = deployment_EP_NETs_1)   
    
    
    #Technology implementation time for EP-NETs technology 2
    def deployment_EP_NETs_2(model, i):
        if model.TIME['EP_NETs_2'][i] == 'NO':
            return model.E[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_28 = pyo.Constraint(i, rule = deployment_EP_NETs_2) 
    
    
    #Technology implementation time for EP-NETs technology 3
    def deployment_EP_NETs_3(model, i):
        if model.TIME['EP_NETs_3'][i] == 'NO':
            return model.F[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_29 = pyo.Constraint(i, rule = deployment_EP_NETs_3)
    
    
    #Technology implementation time for EC-NETs technology 1
    def deployment_EC_NETs_1(model, i):
        if model.TIME['EC_NETs_1'][i] == 'NO':
            return model.G[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_30 = pyo.Constraint(i, rule = deployment_EC_NETs_1)   
    
    
    #Technology implementation time for EC-NETs technology 2
    def deployment_EC_NETs_2(model, i):
        if model.TIME['EC_NETs_2'][i] == 'NO':
            return model.H[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_31 = pyo.Constraint(i, rule = deployment_EC_NETs_2) 
    
    
    #Technology implementation time for EC-NETs technology 3
    def deployment_EC_NETs_3(model, i):
        if model.TIME['EC_NETs_3'][i] == 'NO':
            return model.I[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_32 = pyo.Constraint(i, rule = deployment_EC_NETs_3)
    
    
    #Technology implementation time for solar compensatory energy
    def deployment_solar(model, i):
        if model.TIME['SOLAR'][i] == 'NO':
            return model.J[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_33 = pyo.Constraint(i, rule = deployment_solar)
    
    
    #Technology implementation time for hydropower compensatory energy
    def deployment_hydro(model, i):
        if model.TIME['HYDRO'][i] == 'NO':
            return model.K[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_34 = pyo.Constraint(i, rule = deployment_hydro)
    
    
    #Technology implementation time for biomass compensatory energy
    def deployment_biomass(model, i):
        if model.TIME['BIOMASS'][i] == 'NO':
            return model.L[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_35 = pyo.Constraint(i, rule = deployment_biomass)
    
    
    #Technology implementation time for biogas compensatory energy
    def deployment_biogas(model, i):
        if model.TIME['BIOGAS'][i] == 'NO':
            return model.M[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_36 = pyo.Constraint(i, rule = deployment_biogas)
    
    
    #Technology implementation time for MSW compensatory energy
    def deployment_MSW(model, i):
        if model.TIME['MSW'][i] == 'NO':
            return model.N[i] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_37 = pyo.Constraint(i, rule = deployment_MSW)
    
    
    #Technology implementation time for alternative solid fuel type 1
    def deployment_alt_solid_1(model, i, s):
        if model.TIME['SOLID_1'][i] == 'NO':
            return model.O[i,s] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_38 = pyo.Constraint(i, model.S, rule = deployment_alt_solid_1)

    
    #Technology implementation time for alternative solid fuel type 2
    def deployment_alt_solid_2(model, i, s):
        if model.TIME['SOLID_2'][i] == 'NO':
            return model.P[i,s] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_39 = pyo.Constraint(i, model.S, rule = deployment_alt_solid_2)
    
    
    #Technology implementation time for alternative gas fuel type 1
    def deployment_alt_gas_1(model, i, s):
        if model.TIME['GAS_1'][i] == 'NO':
            return model.Q[i,s] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_40 = pyo.Constraint(i, model.S, rule = deployment_alt_gas_1)

    
    #Technology implementation time for alternative gas fuel type 2
    def deployment_alt_gas_2(model, i, s):
        if model.TIME['GAS_2'][i] == 'NO':
            return model.R[i,s] == 0
        else:
            return pyo.Constraint.Skip
        
    model.Cons_41 = pyo.Constraint(i, model.S, rule = deployment_alt_gas_2)
    
    
    #Big M formulation for EP-NETs technology 1 deployment for period i
    def big_M_EP_NETs_1(model, i):
        return model.EP_NET_1[i] <= model.D[i] * 1000
    
    model.Cons_42 = pyo.Constraint(i, rule = big_M_EP_NETs_1) 
    
    
    #Big M formulation for EP-NETs technology 2 deployment for period i
    def big_M_EP_NETs_2(model, i):
        return model.EP_NET_2[i] <= model.E[i] * 1000
    
    model.Cons_43 = pyo.Constraint(i, rule = big_M_EP_NETs_2) 
    
    
    #Big M formulation for EP-NETs technology 3 deployment for period i
    def big_M_EP_NETs_3(model, i):
        return model.EP_NET_3[i] <= model.F[i] * 1000
    
    model.Cons_44 = pyo.Constraint(i, rule = big_M_EP_NETs_3) 

    
    #Big M formulation for EC-NETs technology 1 deployment for period i
    def big_M_EC_NETs_1(model, i):
        return model.EC_NET_1[i] <= model.G[i] * 1000
    
    model.Cons_45 = pyo.Constraint(i, rule = big_M_EC_NETs_1) 
    
    
    #Big M formulation for EC-NETs technology 2 deployment for period i
    def big_M_EC_NETs_2(model, i):
        return model.EC_NET_2[i] <= model.H[i] * 1000
    
    model.Cons_46 = pyo.Constraint(i, rule = big_M_EC_NETs_2) 
    
    
    #Big M formulation for EC-NETs technology 3 deployment for period i
    def big_M_EC_NETs_3(model, i):
        return model.EC_NET_3[i] <= model.I[i] * 1000
    
    model.Cons_47 = pyo.Constraint(i, rule = big_M_EC_NETs_3) 
    
    
    #Big M formulation for deployment of solar renewable energy for period i
    def big_M_solar(model, i):
        return model.REN_SOLAR[i] <= model.J[i] * 1000
    
    model.Cons_48 = pyo.Constraint(i, rule = big_M_solar) 
    
    
    #Big M formulation for deployment of hydro renewable energy for period i
    def big_M_hydro(model, i):
        return model.REN_HYDRO[i] <= model.K[i] * 1000
    
    model.Cons_49 = pyo.Constraint(i, rule = big_M_hydro) 
    
    
    #Big M formulation for deployment of BM renewable energy for period i
    def big_M_biomass(model, i):
        return model.REN_BM[i] <= model.L[i] * 1000
    
    model.Cons_50 = pyo.Constraint(i, rule = big_M_biomass) 

    
    #Big M formulation for deployment of BG renewable energy for period i
    def big_M_biogas(model, i):
        return model.REN_BG[i] <= model.M[i] * 1000
    
    model.Cons_51 = pyo.Constraint(i, rule = big_M_biogas) 
    
    
    #Big M formulation for deployment of MSW renewable energy for period i
    def big_M_MSW(model, i):
        return model.REN_MSW[i] <= model.N[i] * 1000
    
    model.Cons_52 = pyo.Constraint(i, rule = big_M_MSW) 
    
    
    #Big M formulation for deployment of alternative solid fuel type 1 for period i
    def big_M_alt_solid_1(model, i, s):
        return model.solid_1[i,s] <= model.O[i,s] * model.plant[s]['UB']
    
    model.Cons_53 = pyo.Constraint(i, model.S, rule = big_M_alt_solid_1)
    
    
    #Big M formulation for deployment of alternative solid fuel type 2 for period i
    def big_M_alt_solid_2(model, i, s):
        return model.solid_2[i,s] <= model.P[i,s] * model.plant[s]['UB']
    
    model.Cons_54 = pyo.Constraint(i, model.S, rule = big_M_alt_solid_2)
    
    
    #Big M formulation for deployment of alternative gas fuel type 1 for period i
    def big_M_alt_gas_1(model, i, s):
        return model.gas_1[i,s] <= model.Q[i,s] * model.plant[s]['UB']
    
    model.Cons_55 = pyo.Constraint(i, model.S, rule = big_M_alt_gas_1)
    
    
    #Big M formulation for deployment of alternative gas fuel type 2 for period i
    def big_M_alt_gas_2(model, i, s):
        return model.gas_2[i,s] <= model.R[i,s] * model.plant[s]['UB']
    
    model.Cons_56 = pyo.Constraint(i, model.S, rule = big_M_alt_gas_2)

    
    #The deployment of solar renewable energy at later periods should at least match the deployment in the previous period
    def solar_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.REN_SOLAR[i+1] >= model.REN_SOLAR[i]
        
    model.Cons_57 = pyo.Constraint(i, model.S, rule = solar_time_constraint)
    

    #The deployment of hydro renewable energy at later periods should at least match the deployment in the previous period
    def hydro_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.REN_HYDRO[i+1] >= model.REN_HYDRO[i]
        
    model.Cons_58 = pyo.Constraint(i, model.S, rule = hydro_time_constraint)
    
    
    #The deployment of solar renewable energy at later periods should at least match the deployment in the previous period
    def biomass_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.REN_BM[i+1] >= model.REN_BM[i]
        
    model.Cons_59 = pyo.Constraint(i, model.S, rule = biomass_time_constraint)

    
    #The deployment of hydro renewable energy at later periods should at least match the deployment in the previous period
    def biogas_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.REN_BG[i+1] >= model.REN_BG[i]
        
    model.Cons_60 = pyo.Constraint(i, model.S, rule = biogas_time_constraint)
    
    
    #The deployment of hydro renewable energy at later periods should at least match the deployment in the previous period
    def MSW_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.REN_MSW[i+1] >= model.REN_MSW[i]
        
    model.Cons_61 = pyo.Constraint(i, model.S, rule = MSW_time_constraint)
    
    
    #The deployment of EP-NETs technology 1 at later periods should at least match the deployment in the previous period
    def EP_NETs_1_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.EP_NET_1[i+1] >= model.EP_NET_1[i]
        
    model.Cons_62 = pyo.Constraint(i, model.S, rule = EP_NETs_1_time_constraint)
    
    
    #The deployment of EP-NETs technology 2 at later periods should at least match the deployment in the previous period
    def EP_NETs_2_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.EP_NET_2[i+1] >= model.EP_NET_2[i]
        
    model.Cons_63 = pyo.Constraint(i, model.S, rule = EP_NETs_2_time_constraint)
    
    
    #The deployment of EP-NETs technology 3 at later periods should at least match the deployment in the previous period
    def EP_NETs_3_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.EP_NET_3[i+1] >= model.EP_NET_3[i]
        
    model.Cons_64 = pyo.Constraint(i, model.S, rule = EP_NETs_3_time_constraint)
    
    
    #The deployment of EC-NETs technology 1 at later periods should at least match the deployment in the previous period
    def EC_NETs_1_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.EC_NET_1[i+1] >= model.EC_NET_1[i]
        
    model.Cons_65 = pyo.Constraint(i, model.S, rule = EC_NETs_1_time_constraint)
    
    
    #The deployment of EC-NETs technology 1 at later periods should at least match the deployment in the previous period
    def EC_NETs_2_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.EC_NET_2[i+1] >= model.EC_NET_2[i]
        
    model.Cons_66 = pyo.Constraint(i, model.S, rule = EC_NETs_2_time_constraint)
    
    
    #The deployment of EC-NETs technology 1 at later periods should at least match the deployment in the previous period
    def EC_NETs_3_time_constraint(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.EC_NET_3[i+1] >= model.EC_NET_3[i]
        
    model.Cons_67 = pyo.Constraint(i, model.S, rule = EC_NETs_3_time_constraint)
    
    
    #Total energy contribution from all energy sources to satisfy the total demand for period i
    def total_energy(model, i, s):
        return sum((model.net_energy[i,s] + model.net_energy_CCS_1[i,s] + model.net_energy_CCS_2[i,s] + model.solid_1[i,s] + model.solid_2[i,s] + model.gas_1[i,s] + model.gas_2[i,s]) for s in model.S) + model.REN_SOLAR[i] + model.REN_HYDRO[i] + model.REN_BM[i] + model.REN_BG[i] + model.REN_MSW[i] + model.EP_NET_1[i] + model.EP_NET_2[i] + model.EP_NET_3[i] == model.EP['Demand'][i] + model.EC_NET_1[i] + model.EC_NET_2[i] + model.EC_NET_3[i]
        
    model.Cons_68 = pyo.Constraint(i, model.S, rule = total_energy)
    
    #Determine the net energy available from power plant s with the deployment of CCS technologies 1 & 2 for period i
    def CCS_net_energy(model, i, s):
        return model.net_energy_CCS_1[i,s] + model.net_energy_CCS_2[i,s] == model.net_energy_CCS[i,s]
        
    model.Cons_69 = pyo.Constraint(i, model.S, rule = CCS_net_energy)  
    
    #Constraints below are related to carbon trading model
    #This constraint forbids the non-generation plants to participate in CT
    def Emissions_unit_cost_value(model, i, s):
        if CT == 'ON':
            return model.emissions_unit_cost_new[i,s] == model.A[i,s] * (model.Emissions_unit_cost[s][i])
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_70 = pyo.Constraint(i, model.S, rule = Emissions_unit_cost_value)
    
    
    #This constraint puts upper limit to the emissions bought from the government by each plant s through carbon trading for period i
    def CT_emissions_UL (model, i, s):
        if CT == 'ON':
            return model.CT_emissions[i,s] <= model.plant[s]['UB'] * model.plant[s]['CI'] * model.A[i,s]
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_71 = pyo.Constraint(i, model.S, rule = CT_emissions_UL)
    
    
    #This constraint puts lower limit to the emissions bought from the government by each plant s through carbon trading for period i
    def CT_emissions_LL (model, i, s):
        if CT == 'ON':
            return model.CT_emissions[i,s] >= model.plant[s]['LB'] * model.plant[s]['CI'] * model.A[i,s]
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_72 = pyo.Constraint(i, model.S, rule = CT_emissions_LL)
    
    
    #This constraint calculates net emissions from a plant s for period i
    def CO2_load_plant(model, i, s):
        return ((model.net_energy[i,s] * model.plant[s]['CI']) + (model.net_energy_CCS_1[i,s] * model.plant[s]['CI'] * ((1 - model.CCS_data['RR_1'][i]) / (1 - model.CCS_data['X_1'][i]))) + (model.net_energy_CCS_2[i,s] * model.plant[s]['CI'] * ((1 - model.CCS_data['RR_2'][i]) / (1 - model.CCS_data['X_2'][i]))) + (model.solid_1[i,s] * model.SLD_CI['SOLID_1'][i]) + (model.solid_2[i,s] * model.SLD_CI['SOLID_2'][i]) + (model.gas_1[i,s] * model.GAS_CI['GAS_1'][i]) + (model.gas_2[i,s] * model.GAS_CI['GAS_2'][i])) == model.net_plant_emissions[i,s] 
    
    model.Cons_73 = pyo.Constraint(i, model.S, rule = CO2_load_plant)
    
    
    #This constraint calculates the emissions avaiable for trading between plants by each plant s for period i
    def CT_emissions_tradable (model, i, s):
        if CT == 'ON':
            return ((model.ET_plant_new [i,s] + model.CT_emissions[i,s]) - (model.net_plant_emissions[i,s])) == model.ET_plant [i,s]
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_74 = pyo.Constraint(i, model.S, rule = CT_emissions_tradable)
    
    
    #This constraint balances the emissions for plant s in period i
    def CT_plant_emissions_balance (model, i, s):
        if CT == 'ON':
            return model.ET_plant [i,s] == model.ET_plant_final [i,s] + model.Emissions_left [i,s]
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_75 = pyo.Constraint(i, model.S, rule = CT_plant_emissions_balance)
    
    
    #This constraint calculates balance emissions for all periods (except first one)
    def CT_rem_balance_emissions (model, i, s):
        if CT == 'ON':
            return model.CT_balance_emissions [i] == sum(model.Emissions_left [i,s] for s in model.S)
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_76 = pyo.Constraint(i, model.S, rule = CT_rem_balance_emissions)
    
    
    #This constraint calculates balance emissions for all periods (except first one)
    def CT_trade_within_plants (model, i, s):
        if CT == 'ON':
            return sum(model.ET_plant_final[i,s] for s in model.S) == 0
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_77 = pyo.Constraint(i, model.S, rule = CT_trade_within_plants)
    
    
    #This constraint quantifies the emissions avaible for a plant to be traded within a period
    def CT_plant_emissions_value (model, i, s):
        if CT == 'ON':
            return model.ET_plant_final [i,s] == model.ET_plant_final[i,s] * model.A[i,s]
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_78 = pyo.Constraint(i, model.S, rule = CT_plant_emissions_value)
    
    
    #This constraint enables puts an upper limit on emissions trading between plants
    def CT_plant_UL_final (model, i, s):
        if CT == 'ON':
            return model.ET_plant_final [i,s] <= (CT_Percentage_Emissions) * (model.energy[i,s] * (model.plant[s]['CI']))
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_79 = pyo.Constraint(i, model.S, rule = CT_plant_UL_final)
    
    
    #This constraint enables puts a lower limit on emissions trading between plants
    def CT_plant_LL_final (model, i, s):
        if CT == 'ON':
            return model.ET_plant_final [i,s] >= (-1) * (CT_Percentage_Emissions) * (model.energy[i,s] * (model.plant[s]['CI']))
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_80 = pyo.Constraint(i, model.S, rule = CT_plant_LL_final)
    
    
    #This constraint puts limit to balance emissions traded in next period
    def CT_balance_emissions_trading (model, i, s):
        if CT == 'ON':
            if i == 1:
                return sum(model.ET_plant_new [i,s] for s in model.S) <= EB
            else:
                return sum(model.ET_plant_new [i,s] for s in model.S) <= model.CT_balance_emissions[i-1]
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_81 = pyo.Constraint(i, model.S, rule = CT_balance_emissions_trading)
    
    
    #This constraint enables puts an upper limit on emissions between periods
    def CT_plant_new_UL (model, i, s):
        if CT == 'ON':
            return model.ET_plant_new [i,s] <= (model.plant[s]['UB'] * model.plant[s]['CI'] * model.A[i,s])
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_82 = pyo.Constraint(i, model.S, rule = CT_plant_new_UL)
    
    
    #This constraint puts a budget cap on carbon trading
    def CT_budget_cap (model, i, s):
        if CT == 'ON':
            return ((sum((model.CT_emissions[i,s] * model.emissions_unit_cost_new[i,s]) for s in model.S)) + (model.EP['CT Additional Budget'][i])) >= ((-1) * (sum((model.Emissions_left [i,s] * model.emissions_unit_cost_new[i,s]) for s in model.S)))
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_83 = pyo.Constraint(i, model.S, rule = CT_budget_cap)
    
    
    #This constraint enables trading within different periods
    def CT_periods (model, i, s):
        if CT == 'ON':
            return sum(model.ET_plant_new[i,s] for s in model.S) + sum(model.ET_plant_final[i,s] for s in model.S) + sum(model.CT_emissions[i,s] for s in model.S) == (sum((model.energy[i,s] * (model.plant[s]['CI'])) for s in model.S))
        elif CT == 'OFF':
            return pyo.Constraint.Skip
    
    model.Cons_84 = pyo.Constraint(i, model.S, rule = CT_periods)
    
    
    #This constraint calculates the cost paid/earned by each plant s to buy or sell carbon emissions among each other for period i
    def CT_cost (model, i, s):
        if CT == 'ON':
            return model.CT_cost[i,s] == model.emissions_unit_cost_new[i,s] * (model.CT_emissions[i,s] + model.ET_plant_new [i,s] + model.ET_plant_final [i,s] + model.Emissions_left [i,s])
        elif CT == 'OFF':
            return pyo.Constraint.Skip
        
    model.Cons_85 = pyo.Constraint(i, model.S, rule = CT_cost)
    
   
    #This constraint defines the earning of a power plant s by selling stored emissions as a result of implementing CCS for period i
    def CO2_selling(model, i, s):
        return ((((model.net_energy_CCS[i,s])*(model.plant[s]['CI'])) - ((model.net_energy_CCS_1[i,s]) * (model.plant[s]['CI'] * (1 - model.CCS_data['RR_1'][i]) / (1 - model.CCS_data['X_1'][i]))) - ((model.net_energy_CCS_2[i,s]) * (model.plant[s]['CI'] * (1 - model.CCS_data['RR_2'][i]) / (1 - model.CCS_data['X_2'][i])))) * (model.EP['Carbon Emissions Price'][i])) == (model.CCS_selling_price[i,s]) 
        
    model.Cons_86 = pyo.Constraint(i, model.S, rule = CO2_selling)
    
    
    #This constraint limits the stored emissions as a result of CCS for period i
    def Stored_emissions_limit(model, i, s):
        return model.EP['Carbon Emissions Requirement'][i] >= sum((((model.net_energy_CCS_1[i,s]) * (model.plant[s]['CI'] * (1 - model.CCS_data['RR_1'][i]) / (1 - model.CCS_data['X_1'][i]))) + ((model.net_energy_CCS_2[i,s])*(model.plant[s]['CI'] * (1 - model.CCS_data['RR_2'][i]) / (1 - model.CCS_data['X_2'][i])))) for s in model.S)

    model.Cons_87 = pyo.Constraint(i, model.S, rule = Stored_emissions_limit)
    
    
    #Determining the total earning of a power plant s for period i
    def Plant_earning(model, i, s):
        if 'SOLAR' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Solar)'][i]))
        elif 'WIND' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Wind)'][i]))
        elif 'HYDRO' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Hydro)'][i]))
        elif 'BIOGAS' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Biogas)'][i]))
        elif 'BIOMASS' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Biomass)'][i]))
        elif 'MSW' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (MSW)'][i]))
        elif 'NUCLEAR' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Nuclear)'][i]))
        elif 'NG' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (NG)'][i]))
        elif 'OIL' in model.plant[s].values():
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Oil)'][i]))
        else:
            return model.Plant_earning[i,s] == (model.energy[i,s] * (model.Electricity_price[i,s]) * (1 - model.EP['Tax (Coal)'][i]))
    
    model.Cons_88 = pyo.Constraint(i, model.S, rule = Plant_earning)
    
    
    #This constraint puts constraint to electricity prices
    def Electricity_Prices(model, i, s):
        if i == numperiods - 1:
            return pyo.Constraint.Skip
        else:
            return model.Electricity_price[i+1,s] >= (model.Electricity_price[i,s]) * ((1)+(Elec_Price_Inflation))
        
    model.Cons_89 = pyo.Constraint(i, model.S, rule = Electricity_Prices)
    
    
    #This constraint defines the initial electricity price for first period
    def Electricity_Price_P1(model, i, s):
        if i == 1:
            return model.Electricity_price[i,s] == (Initial_Elec_Price)
        else:
            return model.Electricity_price[i,s] <= (Max_Elec_Price)
        
    model.Cons_90 = pyo.Constraint(i, model.S, rule = Electricity_Price_P1)
    
    
    #The total CO2 load contribution from all energy sources must satisfy most the CO2 emission limit in period i
    def total_CO2_load(model, i, s):
        return (sum((model.net_energy[i,s] * model.plant[s]['CI']) + (model.net_energy_CCS_1[i,s] * model.plant[s]['CI'] * (1 - model.CCS_data['RR_1'][i]) / (1 - model.CCS_data['X_1'][i])) + (model.net_energy_CCS_2[i,s] * model.plant[s]['CI'] * (1 - model.CCS_data['RR_2'][i]) / (1 - model.CCS_data['X_2'][i])) 
        + (model.solid_1[i,s] * model.SLD_CI['SOLID_1'][i]) + (model.solid_2[i,s] * model.SLD_CI['SOLID_2'][i]) 
        + (model.gas_1[i,s] * model.GAS_CI['GAS_1'][i]) + (model.gas_2[i,s] * model.GAS_CI['GAS_2'][i]) for s in model.S) 
        + (model.EC_NET_1[i] * model.NET_CI['EC_NETs_1'][i])
        + (model.EC_NET_2[i] * model.NET_CI['EC_NETs_2'][i])
        + (model.EC_NET_3[i] * model.NET_CI['EC_NETs_3'][i])
        + (model.EP_NET_1[i] * model.NET_CI['EP_NETs_1'][i]) 
        + (model.EP_NET_2[i] * model.NET_CI['EP_NETs_2'][i])
        + (model.EP_NET_3[i] * model.NET_CI['EP_NETs_3'][i])
        + (model.REN_SOLAR[i] * model.REN_CI['SOLAR'][i])
        + (model.REN_HYDRO[i] * model.REN_CI['HYDRO'][i]) 
        + (model.REN_BM[i] * model.REN_CI['BIOMASS'][i])
        + (model.REN_BG[i] * model.REN_CI['BIOGAS'][i]) 
        + (model.REN_MSW[i] * model.REN_CI['MSW'][i]) == model.new_emission[i])

    model.Cons_91 = pyo.Constraint(i, model.S, rule = total_CO2_load)
       
    
    #Determining the cumulative total fuel and annualised capital cost for a power plant s for period i
    def energy_cost(model, i, s):
        if 'SOLAR' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['SOLAR'][i]) + (AFF * model.CPX_1['SOLAR'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['SOLAR'][i])
        elif 'WIND' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['WIND'][i]) + (AFF * model.CPX_1['WIND'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['WIND'][i])
        elif 'HYDRO' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['HYDRO'][i]) + (AFF * model.CPX_1['HYDRO'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['HYDRO'][i])
        elif 'BIOGAS' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['BIOGAS'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['BIOGAS'][i])
        elif 'BIOMASS' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['BIOMASS'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['BIOMASS'][i])
        elif 'MSW' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['MSW'][i]) + (AFF * model.CPX_1['MSW'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['MSW'][i])
        elif 'NUCLEAR' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['NUCLEAR'][i]) + (AFF * model.CPX_1['NUCLEAR'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['NUCLEAR'][i])
        elif 'NG' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['NG'][i]) + (AFF * model.CPX_1['NG'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['NG'][i])
        elif 'OIL' in model.plant[s].values():
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['OIL'][i]) + (AFF * model.CPX_1['OIL'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['OIL'][i])
        else:
            return model.energy_cost[i,s] == (model.net_energy[i,s] * model.fuel['COAL'][i]) + (AFF * model.CPX_1['COAL'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['COAL'][i])
    
    model.Cons_92 = pyo.Constraint(i, model.S, rule = energy_cost)  
    
    
    #The summation of cost for each power plant s should equal to the total cost of each period i
    def Plant_net_cost(model, i, s):
        if CT == 'OFF':
            return ((model.Plant_earning[i,s]) - ((model.net_energy_CCS_1[i,s] * model.CCS_data['Cost_CCS_1'][i]) + (model.net_energy_CCS_2[i,s] * model.CCS_data['Cost_CCS_2'][i]) 
            + (AFF * model.CCS_data['FX_Cost_CCS_1'][i] * model.B[i,s]) + (AFF * model.CCS_data['FX_Cost_CCS_2'][i] * model.C[i,s])
            + (model.solid_1[i,s] * model.SLD_COST['SOLID_1'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.O[i,s]) + (model.solid_2[i,s] * model.SLD_COST['SOLID_2'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.P[i,s])
            + (model.gas_1[i,s] * model.GAS_COST['GAS_1'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.Q[i,s]) + (model.gas_2[i,s] * model.GAS_COST['GAS_2'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.R[i,s])
            + (model.energy_cost[i,s]) - (model.CCS_selling_price[i,s]))) == (model.Plant_cost[i,s])
        
        elif CT == 'ON':
            return (model.Plant_earning[i,s]) - ((model.net_energy_CCS_1[i,s] * model.CCS_data['Cost_CCS_1'][i]) + (model.net_energy_CCS_2[i,s] * model.CCS_data['Cost_CCS_2'][i]) 
            + (AFF * model.CCS_data['FX_Cost_CCS_1'][i] * model.B[i,s]) + (AFF * model.CCS_data['FX_Cost_CCS_2'][i] * model.C[i,s])
            + (model.solid_1[i,s] * model.SLD_COST['SOLID_1'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.O[i,s]) + (model.solid_2[i,s] * model.SLD_COST['SOLID_2'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.P[i,s])
            + (model.gas_1[i,s] * model.GAS_COST['GAS_1'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.Q[i,s]) + (model.gas_2[i,s] * model.GAS_COST['GAS_2'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.R[i,s])
            + (model.energy_cost[i,s]) + (model.CT_cost[i,s]) - (model.CCS_selling_price[i,s])) == (model.Plant_cost[i,s])
   
    model.Cons_93 = pyo.Constraint(i, model.S, rule = Plant_net_cost)

    
    #Determining the total fuel cost for a power plant s for period i
    def fuel_cost(model, i, s):
        if 'SOLAR' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['SOLAR'][i])
        elif 'WIND' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['WIND'][i])
        elif 'HYDRO' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['HYDRO'][i])
        elif 'BIOGAS' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['BIOGAS'][i])
        elif 'BIOMASS' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['BIOMASS'][i])
        elif 'MSW' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['MSW'][i])
        elif 'NUCLEAR' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['NUCLEAR'][i])
        elif 'NG' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['NG'][i])
        elif 'OIL' in model.plant[s].values():
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['OIL'][i])
        else:
            return model.fuel_cost[i,s] == (model.net_energy[i,s] * model.fuel['COAL'][i])
    
    model.Cons_94 = pyo.Constraint(i, model.S, rule = fuel_cost) 
    
    
    #Determining the total annualised capital cost for a power plant s for period i
    def annualized_capital_cost(model, i, s):
        if 'SOLAR' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['SOLAR'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['SOLAR'][i])
        elif 'WIND' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['WIND'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['WIND'][i])
        elif 'HYDRO' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['HYDRO'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['HYDRO'][i])
        elif 'BIOGAS' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['BIOGAS'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['BIOGAS'][i])
        elif 'BIOMASS' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['BIOMASS'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['BIOMASS'][i])
        elif 'MSW' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['MSW'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['MSW'][i])
        elif 'NUCLEAR' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['NUCLEAR'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['NUCLEAR'][i])
        elif 'NG' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['NG'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['NG'][i])
        elif 'OIL' in model.plant[s].values():
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['OIL'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['OIL'][i])
        else:
            return model.annualized_capital_cost[i,s] == (AFF * model.CPX_1['COAL'][i] * model.energy[i,s]) + (AFF * model.energy[i,s] * model.CPX_2['COAL'][i])
    
    model.Cons_95 = pyo.Constraint(i, model.S, rule = annualized_capital_cost) 
       
    
    #The summation of cost for each power plant s should equal to the total cost of each period i
    def sum_cost(model, i, s):
        if CT == 'OFF':
            return (sum((model.net_energy_CCS_1[i,s] * model.CCS_data['Cost_CCS_1'][i]) + (model.net_energy_CCS_2[i,s] * model.CCS_data['Cost_CCS_2'][i]) 
            + (AFF * model.CCS_data['FX_Cost_CCS_1'][i] * model.B[i,s]) + (AFF * model.CCS_data['FX_Cost_CCS_2'][i] * model.C[i,s])
            + (model.solid_1[i,s] * model.SLD_COST['SOLID_1'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.O[i,s]) + (model.solid_2[i,s] * model.SLD_COST['SOLID_2'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.P[i,s])
            + (model.gas_1[i,s] * model.GAS_COST['GAS_1'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.Q[i,s]) + (model.gas_2[i,s] * model.GAS_COST['GAS_2'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.R[i,s]) for s in model.S)
            + (model.EC_NET_1[i] * model.NET_COST['EC_NETs_1'][i]) + (AFF * model.CPX_1['EC_NETs_1'][i] * model.G[i]) + (AFF * model.EC_NET_1[i] * model.CPX_2['EC_NETs_1'][i])
            + (model.EC_NET_2[i] * model.NET_COST['EC_NETs_2'][i]) + (AFF * model.CPX_1['EC_NETs_2'][i] * model.H[i]) + (AFF * model.EC_NET_2[i] * model.CPX_2['EC_NETs_2'][i])
            + (model.EC_NET_3[i] * model.NET_COST['EC_NETs_3'][i]) + (AFF * model.CPX_1['EC_NETs_3'][i] * model.I[i]) + (AFF * model.EC_NET_3[i] * model.CPX_2['EC_NETs_3'][i])
            + (model.EP_NET_1[i] * model.NET_COST['EP_NETs_1'][i]) + (AFF * model.CPX_1['EP_NETs_1'][i] * model.D[i]) + (AFF * model.EP_NET_1[i] * model.CPX_2['EP_NETs_1'][i])
            + (model.EP_NET_2[i] * model.NET_COST['EP_NETs_2'][i]) + (AFF * model.CPX_1['EP_NETs_2'][i] * model.E[i]) + (AFF * model.EP_NET_2[i] * model.CPX_2['EP_NETs_2'][i])
            + (model.EP_NET_3[i] * model.NET_COST['EP_NETs_3'][i]) + (AFF * model.CPX_1['EP_NETs_3'][i] * model.F[i]) + (AFF * model.EP_NET_3[i] * model.CPX_2['EP_NETs_3'][i])
            + (model.REN_SOLAR[i] * model.REN_COST['SOLAR'][i]) + (AFF * model.CPX_1['SOLAR'][i] * model.J[i]) + (AFF * model.REN_SOLAR[i] * model.CPX_2['SOLAR'][i])
            + (model.REN_HYDRO[i] * model.REN_COST['HYDRO'][i]) + (AFF * model.CPX_1['HYDRO'][i] * model.K[i]) + (AFF * model.REN_HYDRO[i] * model.CPX_2['HYDRO'][i])
            + (model.REN_BM[i] * model.REN_COST['BIOMASS'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.L[i]) + (AFF * model.REN_BM[i] * model.CPX_2['BIOMASS'][i])
            + (model.REN_BG[i] * model.REN_COST['BIOGAS'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.M[i]) + (AFF * model.REN_BG[i] * model.CPX_2['BIOGAS'][i])
            + (model.REN_MSW[i] * model.REN_COST['MSW'][i]) + (AFF * model.CPX_1['MSW'][i] * model.N[i]) + (AFF * model.REN_MSW[i] * model.CPX_2['MSW'][i])
            + sum(model.energy_cost[i,s] for s in model.S) - sum(model.CCS_selling_price[i,s] for s in model.S) == model.sum_cost[i])
        
        elif CT == 'ON':
            return (sum((model.net_energy_CCS_1[i,s] * model.CCS_data['Cost_CCS_1'][i]) + (model.net_energy_CCS_2[i,s] * model.CCS_data['Cost_CCS_2'][i]) 
            + (AFF * model.CCS_data['FX_Cost_CCS_1'][i] * model.B[i,s]) + (AFF * model.CCS_data['FX_Cost_CCS_2'][i] * model.C[i,s])
            + (model.solid_1[i,s] * model.SLD_COST['SOLID_1'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.O[i,s]) + (model.solid_2[i,s] * model.SLD_COST['SOLID_2'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.P[i,s])
            + (model.gas_1[i,s] * model.GAS_COST['GAS_1'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.Q[i,s]) + (model.gas_2[i,s] * model.GAS_COST['GAS_2'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.R[i,s]) for s in model.S)
            + (model.EC_NET_1[i] * model.NET_COST['EC_NETs_1'][i]) + (AFF * model.CPX_1['EC_NETs_1'][i] * model.G[i]) + (AFF * model.EC_NET_1[i] * model.CPX_2['EC_NETs_1'][i])
            + (model.EC_NET_2[i] * model.NET_COST['EC_NETs_2'][i]) + (AFF * model.CPX_1['EC_NETs_2'][i] * model.H[i]) + (AFF * model.EC_NET_2[i] * model.CPX_2['EC_NETs_2'][i])
            + (model.EC_NET_3[i] * model.NET_COST['EC_NETs_3'][i]) + (AFF * model.CPX_1['EC_NETs_3'][i] * model.I[i]) + (AFF * model.EC_NET_3[i] * model.CPX_2['EC_NETs_3'][i])
            + (model.EP_NET_1[i] * model.NET_COST['EP_NETs_1'][i]) + (AFF * model.CPX_1['EP_NETs_1'][i] * model.D[i]) + (AFF * model.EP_NET_1[i] * model.CPX_2['EP_NETs_1'][i])
            + (model.EP_NET_2[i] * model.NET_COST['EP_NETs_2'][i]) + (AFF * model.CPX_1['EP_NETs_2'][i] * model.E[i]) + (AFF * model.EP_NET_2[i] * model.CPX_2['EP_NETs_2'][i])
            + (model.EP_NET_3[i] * model.NET_COST['EP_NETs_3'][i]) + (AFF * model.CPX_1['EP_NETs_3'][i] * model.F[i]) + (AFF * model.EP_NET_3[i] * model.CPX_2['EP_NETs_3'][i])
            + (model.REN_SOLAR[i] * model.REN_COST['SOLAR'][i]) + (AFF * model.CPX_1['SOLAR'][i] * model.J[i]) + (AFF * model.REN_SOLAR[i] * model.CPX_2['SOLAR'][i])
            + (model.REN_HYDRO[i] * model.REN_COST['HYDRO'][i]) + (AFF * model.CPX_1['HYDRO'][i] * model.K[i]) + (AFF * model.REN_HYDRO[i] * model.CPX_2['HYDRO'][i])
            + (model.REN_BM[i] * model.REN_COST['BIOMASS'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.L[i]) + (AFF * model.REN_BM[i] * model.CPX_2['BIOMASS'][i])
            + (model.REN_BG[i] * model.REN_COST['BIOGAS'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.M[i]) + (AFF * model.REN_BG[i] * model.CPX_2['BIOGAS'][i])
            + (model.REN_MSW[i] * model.REN_COST['MSW'][i]) + (AFF * model.CPX_1['MSW'][i] * model.N[i]) + (AFF * model.REN_MSW[i] * model.CPX_2['MSW'][i])
            + sum(model.energy_cost[i,s] for s in model.S) + sum(model.CT_cost[i,s] for s in model.S) - sum(model.CCS_selling_price[i,s] for s in model.S) == model.sum_cost[i])
   
    model.Cons_96 = pyo.Constraint(i, model.S, rule = sum_cost)
    
    
    def Total_profit (model, i, s):
        return (sum(model.Plant_earning[i,s] for s in model.S)) - (model.sum_cost[i]) == model.total_profit[i]

    model.Cons_97 = pyo.Constraint(i, model.S, rule = Total_profit)
    
    
    #This constraint ensures the net profitability over period i
    def Plant_profit_net (model, i, s):
        if Plant_Profit == 'ON':
            return model.Plant_cost[i,s] >= 0
        elif Plant_Profit == 'OFF':
            return pyo.Constraint.Skip
        
    model.Cons_98 = pyo.Constraint(i, model.S, rule = Plant_profit_net)
    
    
    #For the minimum budget objective function, the total cost is minimised, subject to the satisfaction of the CO2 emission limit
    #For the minimum emission objective function, the total emission is minimised, subject to the available budgetary constraint
    def objective_constraint(model, i):
        if flag == 'min_budget':
            return model.new_emission[i] <= model.EP['Limit'][i]
        elif flag == 'min_emission':
            return model.sum_cost[i] <= model.EP['Budget'][i]
        else:
            return pyo.Constraint.Skip
    
    model.Cons_99 = pyo.Constraint(i, rule = objective_constraint)
    
    #Constraints 100 and 101 are to check the feasibility of the model, and it also performs maximization of profits 
    def objective_constraint(model, i):
        if flag == 'max_profit':
            return model.new_emission[i] <= model.EP['Limit'][i]
        else:
            return pyo.Constraint.Skip
    
    model.Cons_100 = pyo.Constraint(i, rule = objective_constraint)
    
    
    def objective_constraint(model, i):
        if flag == 'max_profit':
            return model.sum_cost[i] <= model.EP['Budget'][i]
        else:
            return pyo.Constraint.Skip
    
    model.Cons_101 = pyo.Constraint(i, rule = objective_constraint)
    
    
    #opt = SolverFactory('octeract-engine', tee = True)
    #results = opt.solve(model)
    
    opt = SolverFactory('gams')
    #sys.exit()
  
    results = opt.solve(model, solver = 'scip', tee = True, add_options = ['option reslim = 36000;', 'option optcr = 0.01;', 'option optca = 0.01;', 'GAMS_MODEL.nodlim = 1000000;'])
    #optca: Absolute optimality criterion for MIPs, and is of the order of 10^-9 by default
    #optcr: This option specifies a relative termination tolerance for use in solving MIP problems, default is 5E-8, set it between 0.01 and 0.05 for faster computation, having a low value for this gives better results
    
    #opt = SolverFactory('gurobi')
    #results = opt.solve(model, tee = True)
    
    print(results)
    #model.pprint()   
    return model

def multiperiod_energy_planning_results(model, i):
    energy_planning = pd.DataFrame()

    for s in model.plant.keys():
        energy_planning.loc[s, 'Fuel'] = model.plant[s]['Fuel']
        energy_planning.loc[s, 'Energy Generation'] = model.A[i,s]()
        energy_planning.loc[s, 'Gross Energy (TWh/y)'] = round(model.energy[i,s](), 3)
        energy_planning.loc[s, 'CO2 Intensity (Mt/TWh)'] = model.plant[s]['CI']
        energy_planning.loc[s, 'CCS_1 CI'] = round(model.CI_RET_1[i,s](), 3)
        energy_planning.loc[s, 'CCS_2 CI'] = round(model.CI_RET_2[i,s](), 3)
        energy_planning.loc[s, 'CCS_1 Selection'] = model.B[i,s]()
        energy_planning.loc[s, 'CCS_1 Ret (TWh/y)'] = round(model.CCS_1[i,s](), 3)
        energy_planning.loc[s, 'CCS_2 Selection'] = model.C[i,s]()         
        energy_planning.loc[s, 'CCS_2 Ret (TWh/y)'] = round(model.CCS_2[i,s](), 3)       
        energy_planning.loc[s, 'Net Energy wo CCS'] = round(model.net_energy[i,s](), 3)
        energy_planning.loc[s, 'Net Energy w CCS_1'] = round(model.net_energy_CCS_1[i,s](), 3)
        energy_planning.loc[s, 'Net Energy w CCS_2'] = round(model.net_energy_CCS_2[i,s](), 3)
        energy_planning.loc[s, 'Solid_1 Selection'] = model.O[i,s]()        
        energy_planning.loc[s, 'SOLID_1 (TWh/y)'] = round(model.solid_1[i,s](), 3)
        energy_planning.loc[s, 'Solid_2 Selection'] = model.P[i,s]()
        energy_planning.loc[s, 'SOLID_2 (TWh/y)'] = round(model.solid_2[i,s](), 3)
        energy_planning.loc[s, 'Gas_1 Selection'] = model.Q[i,s]() 
        energy_planning.loc[s, 'GAS_1 (TWh/y)'] = round(model.gas_1[i,s](), 3)
        energy_planning.loc[s, 'Gas_2 Selection'] = model.R[i,s]()
        energy_planning.loc[s, 'GAS_2 (TWh/y)'] = round(model.gas_2[i,s](), 3)
        energy_planning.loc[s, 'Net Energy (TWh/y)'] = round(model.net_energy[i,s]() + model.net_energy_CCS_1[i,s]() + model.net_energy_CCS_2[i,s]() + model.solid_1[i,s]() + model.solid_2[i,s]() + model.gas_1[i,s]() + model.gas_2[i,s](), 3)
        energy_planning.loc[s, 'CO2 Load (Mt/y)'] = round(model.net_plant_emissions[i,s](), 3)
        #Below lines are commented by the if/else statement if CT is turned off
        if CT == 'ON':
            energy_planning.loc[s, 'CT Emissions Price (mil USD / Mt)'] = round(model.emissions_unit_cost_new[i,s](), 3)
            energy_planning.loc[s, 'Emissions bought from previous period (Mt/y)'] = round(model.ET_plant_new[i,s](), 3)
            energy_planning.loc[s, 'Emission Rights Bought by Each Plant from government (Mt/y)'] = round(model.CT_emissions [i,s](), 3)
            energy_planning.loc[s, 'Emissions rights available for trading within companies (Mt/y)'] = round(model.ET_plant[i,s](), 3)
            energy_planning.loc[s, 'Emissions traded within a company (Mt/y)'] = round(model.ET_plant_final[i,s](), 3)
            energy_planning.loc[s, 'Net Emissions traded within a company (Mt/y)'] = round(((model.ET_plant[i,s]()) - (model.ET_plant_final[i,s]())), 3)
            energy_planning.loc[s, 'CT Cost Paid to government to buy emissions from previous period (mil USD / y)'] = round((model.ET_plant_new[i,s]() * model.emissions_unit_cost_new[i,s]()), 3)
            energy_planning.loc[s, 'CT Cost Paid by Plant to government (mil USD / y)'] = round((model.CT_emissions [i,s]() * model.emissions_unit_cost_new[i,s]()), 3)
            energy_planning.loc[s, 'CT Cost Paid/earned by each plant as a result of trading with each other (mil USD / y)'] = round(model.emissions_unit_cost_new[i,s]() * model.ET_plant_final [i,s](), 3)
            energy_planning.loc[s, 'Balance Emissions to be traded in next period'] = round(model.Emissions_left[i,s](), 3)
            energy_planning.loc[s, 'Price paid/earned for balance emissions (mil USD / y)'] = round((model.Emissions_left[i,s]() * model.emissions_unit_cost_new[i,s]()), 3)        
        else:
            pass
        #Above lines are commented by the if/else statement if CT is turned off
        energy_planning.loc[s, 'Cumulative total fuel and annualized capital cost (mil USD/y)'] = round(model.energy_cost[i,s](), 3)
        energy_planning.loc[s, 'Annualized capital cost (mil USD/y)'] = round(model.annualized_capital_cost[i,s](), 3)
        energy_planning.loc[s, 'Fuel cost (mil USD/y)'] = round(model.fuel_cost[i,s](), 3)
        energy_planning.loc[s, 'Cost contribution of CCS_1 (mil USD/y)'] = round((model.net_energy_CCS_1[i,s]() * model.CCS_data['Cost_CCS_1'][i]) + (AFF * model.CCS_data['FX_Cost_CCS_1'][i] * model.B[i,s]()), 3)
        energy_planning.loc[s, 'Cost contribution of CCS_2 (mil USD/y)'] = round((model.net_energy_CCS_2[i,s]() * model.CCS_data['Cost_CCS_2'][i]) + (AFF * model.CCS_data['FX_Cost_CCS_2'][i] * model.C[i,s]()), 3)
        energy_planning.loc[s, 'Cost contribution of Solid_1 (mil USD/y)'] = round((model.solid_1[i,s]() * model.SLD_COST['SOLID_1'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.O[i,s]()), 3)
        energy_planning.loc[s, 'Cost contribution of Solid_2 (mil USD/y)'] = round((model.solid_2[i,s]() * model.SLD_COST['SOLID_2'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.P[i,s]()), 3)
        energy_planning.loc[s, 'Cost contribution of Gas_1 (mil USD/y)'] = round((model.gas_1[i,s]() * model.GAS_COST['GAS_1'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.Q[i,s]()), 3)
        energy_planning.loc[s, 'Cost contribution of Gas_2 (mil USD/y)'] = round((model.gas_2[i,s]() * model.GAS_COST['GAS_2'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.R[i,s]()), 3)
        energy_planning.loc[s, 'CO2 Sold by Plant (Mt/y)'] = round((((model.net_energy_CCS[i,s]())*(model.plant[s]['CI'])) - ((model.net_energy_CCS_1[i,s]())*(model.CI_RET_1[i,s]())) - ((model.net_energy_CCS_2[i,s]())*(model.CI_RET_2[i,s]()))), 3)
        energy_planning.loc[s, 'Earning by CO2 selling (mil USD / y)'] = round(model.CCS_selling_price[i,s](), 3)
        energy_planning.loc[s, 'Electricty Prices (mil USD / TWh)'] = round(model.Electricity_price[i,s]() * model.A[i,s](), 3)
        energy_planning.loc[s, 'Earning by selling electricty (mil USD / y)'] = round(model.Plant_earning[i,s](), 3)
        energy_planning.loc[s, 'Plant Profit (mil USD/y)'] = round(model.Plant_cost[i,s](), 3)
        energy_planning.loc[s, 'Total Cost (mil USD/y)'] = round((model.Plant_earning[i,s]()) - (model.Plant_cost[i,s]()), 3)
        
    energy_planning.loc['EP_NET_1', 'Fuel'] = 'EP_NET_1'
    energy_planning.loc['EP_NET_2', 'Fuel'] = 'EP_NET_2'
    energy_planning.loc['EP_NET_3', 'Fuel'] = 'EP_NET_3'
    energy_planning.loc['EC_NET_1', 'Fuel'] = 'EC_NET_1'
    energy_planning.loc['EC_NET_2', 'Fuel'] = 'EC_NET_2' 
    energy_planning.loc['EC_NET_3', 'Fuel'] = 'EC_NET_3' 
    energy_planning.loc['EP_NET_1', 'Energy Generation'] = model.D[i]()
    energy_planning.loc['EP_NET_2', 'Energy Generation'] = model.E[i]()
    energy_planning.loc['EP_NET_3', 'Energy Generation'] = model.F[i]()
    energy_planning.loc['EC_NET_1', 'Energy Generation'] = model.G[i]()
    energy_planning.loc['EC_NET_2', 'Energy Generation'] = model.H[i]()
    energy_planning.loc['EC_NET_3', 'Energy Generation'] = model.I[i]()
    energy_planning.loc['EP_NET_1', 'CO2 Intensity (Mt/TWh)'] = model.NET_CI['EP_NETs_1'][i]
    energy_planning.loc['EP_NET_2', 'CO2 Intensity (Mt/TWh)'] = model.NET_CI['EP_NETs_2'][i]
    energy_planning.loc['EP_NET_3', 'CO2 Intensity (Mt/TWh)'] = model.NET_CI['EP_NETs_3'][i]
    energy_planning.loc['EC_NET_1', 'CO2 Intensity (Mt/TWh)'] = model.NET_CI['EC_NETs_1'][i]
    energy_planning.loc['EC_NET_2', 'CO2 Intensity (Mt/TWh)'] = model.NET_CI['EC_NETs_2'][i]
    energy_planning.loc['EC_NET_3', 'CO2 Intensity (Mt/TWh)'] = model.NET_CI['EC_NETs_3'][i]        
    energy_planning.loc['EP_NET_1', 'Net Energy (TWh/y)'] = round(model.EP_NET_1[i](), 3)
    energy_planning.loc['EP_NET_2', 'Net Energy (TWh/y)'] = round(model.EP_NET_2[i](), 3)
    energy_planning.loc['EP_NET_3', 'Net Energy (TWh/y)'] = round(model.EP_NET_3[i](), 3)
    energy_planning.loc['EC_NET_1', 'Net Energy (TWh/y)'] = round(model.EC_NET_1[i](), 3)
    energy_planning.loc['EC_NET_2', 'Net Energy (TWh/y)'] = round(model.EC_NET_2[i](), 3)
    energy_planning.loc['EC_NET_3', 'Net Energy (TWh/y)'] = round(model.EC_NET_3[i](), 3)
    energy_planning.loc['EP_NET_1', 'CO2 Load (Mt/y)'] = round(model.EP_NET_1[i]() * model.NET_CI['EP_NETs_1'][i], 3)
    energy_planning.loc['EP_NET_2', 'CO2 Load (Mt/y)'] = round(model.EP_NET_2[i]() * model.NET_CI['EP_NETs_2'][i], 3)
    energy_planning.loc['EP_NET_3', 'CO2 Load (Mt/y)'] = round(model.EP_NET_3[i]() * model.NET_CI['EP_NETs_3'][i], 3)
    energy_planning.loc['EC_NET_1', 'CO2 Load (Mt/y)'] = round(model.EC_NET_1[i]() * model.NET_CI['EC_NETs_1'][i], 3)
    energy_planning.loc['EC_NET_2', 'CO2 Load (Mt/y)'] = round(model.EC_NET_2[i]() * model.NET_CI['EC_NETs_2'][i], 3)
    energy_planning.loc['EC_NET_3', 'CO2 Load (Mt/y)'] = round(model.EC_NET_3[i]() * model.NET_CI['EC_NETs_3'][i], 3)
    energy_planning.loc['SOLAR', 'Fuel'] = 'SOLAR'  
    energy_planning.loc['HYDRO', 'Fuel'] = 'HYDRO'
    energy_planning.loc['BIOMASS', 'Fuel'] = 'BIOMASS'  
    energy_planning.loc['BIOGAS', 'Fuel'] = 'BIOGAS'
    energy_planning.loc['MSW', 'Fuel'] = 'MSW'
    energy_planning.loc['SOLAR', 'Energy Generation'] = model.J[i]() 
    energy_planning.loc['HYDRO', 'Energy Generation'] = model.K[i]() 
    energy_planning.loc['BIOMASS', 'Energy Generation'] = model.L[i]() 
    energy_planning.loc['BIOGAS', 'Energy Generation'] = model.M[i]() 
    energy_planning.loc['MSW', 'Energy Generation'] = model.N[i]() 
    energy_planning.loc['SOLAR', 'CO2 Intensity (Mt/TWh)'] = model.REN_CI['SOLAR'][i]  
    energy_planning.loc['HYDRO', 'CO2 Intensity (Mt/TWh)'] = model.REN_CI['HYDRO'][i]
    energy_planning.loc['BIOMASS', 'CO2 Intensity (Mt/TWh)'] = model.REN_CI['BIOMASS'][i]  
    energy_planning.loc['BIOGAS', 'CO2 Intensity (Mt/TWh)'] = model.REN_CI['BIOGAS'][i]
    energy_planning.loc['MSW', 'CO2 Intensity (Mt/TWh)'] = model.REN_CI['MSW'][i]
    energy_planning.loc['SOLAR', 'Net Energy (TWh/y)'] = round(model.REN_SOLAR[i](), 3) 
    energy_planning.loc['SOLAR', 'CO2 Load (Mt/y)'] = round(model.REN_SOLAR[i]() * model.REN_CI['SOLAR'][i], 3)
    energy_planning.loc['HYDRO', 'Net Energy (TWh/y)'] = round(model.REN_HYDRO[i](), 3) 
    energy_planning.loc['HYDRO', 'CO2 Load (Mt/y)'] = round(model.REN_HYDRO[i]() * model.REN_CI['HYDRO'][i], 3)
    energy_planning.loc['BIOMASS', 'Net Energy (TWh/y)'] = round(model.REN_BM[i](), 3) 
    energy_planning.loc['BIOMASS', 'CO2 Load (Mt/y)'] = round(model.REN_BM[i]() * model.REN_CI['BIOMASS'][i], 3)
    energy_planning.loc['BIOGAS', 'Net Energy (TWh/y)'] = round(model.REN_BG[i](), 3) 
    energy_planning.loc['BIOGAS', 'CO2 Load (Mt/y)'] = round(model.REN_BG[i]() * model.REN_CI['BIOGAS'][i], 3)
    energy_planning.loc['MSW', 'Net Energy (TWh/y)'] = round(model.REN_MSW[i](), 3) 
    energy_planning.loc['MSW', 'CO2 Load (Mt/y)'] = round(model.REN_MSW[i]() * model.REN_CI['MSW'][i], 3)
    energy_planning.loc['TOTAL', 'CO2 Load (Mt/y)'] = round(model.new_emission[i](), 3)
    energy_planning.loc['EC_NET_1', 'Total Cost (mil USD/y)'] = round(sum(((model.EC_NET_1[i]() * model.NET_COST['EC_NETs_1'][i]) + (AFF * model.CPX_1['EC_NETs_1'][i] * model.G[i]()) + (AFF * model.EC_NET_1[i]() * model.CPX_2['EC_NETs_1'][i])) for s in model.S), 3)
    energy_planning.loc['EC_NET_2', 'Total Cost (mil USD/y)'] = round(sum(((model.EC_NET_2[i]() * model.NET_COST['EC_NETs_2'][i]) + (AFF * model.CPX_1['EC_NETs_2'][i] * model.H[i]()) + (AFF * model.EC_NET_2[i]() * model.CPX_2['EC_NETs_2'][i])) for s in model.S), 3)
    energy_planning.loc['EC_NET_3', 'Total Cost (mil USD/y)'] = round(sum(((model.EC_NET_3[i]() * model.NET_COST['EC_NETs_3'][i]) + (AFF * model.CPX_1['EC_NETs_3'][i] * model.I[i]()) + (AFF * model.EC_NET_3[i]() * model.CPX_2['EC_NETs_3'][i])) for s in model.S), 3)
    energy_planning.loc['EP_NET_1', 'Total Cost (mil USD/y)'] = round(sum(((model.EP_NET_1[i]() * model.NET_COST['EP_NETs_1'][i]) + (AFF * model.CPX_1['EP_NETs_1'][i] * model.D[i]()) + (AFF * model.EP_NET_1[i]() * model.CPX_2['EP_NETs_1'][i])) for s in model.S), 3)
    energy_planning.loc['EP_NET_2', 'Total Cost (mil USD/y)'] = round(sum(((model.EP_NET_2[i]() * model.NET_COST['EP_NETs_2'][i]) + (AFF * model.CPX_1['EP_NETs_2'][i] * model.E[i]()) + (AFF * model.EP_NET_2[i]() * model.CPX_2['EP_NETs_2'][i])) for s in model.S), 3)
    energy_planning.loc['EP_NET_3', 'Total Cost (mil USD/y)'] = round(sum(((model.EP_NET_3[i]() * model.NET_COST['EP_NETs_3'][i]) + (AFF * model.CPX_1['EP_NETs_3'][i] * model.F[i]()) + (AFF * model.EP_NET_3[i]() * model.CPX_2['EP_NETs_3'][i])) for s in model.S), 3)
    energy_planning.loc['SOLAR', 'Total Cost (mil USD/y)'] = round(sum(((model.REN_SOLAR[i]() * model.REN_COST['SOLAR'][i]) + (AFF * model.CPX_1['SOLAR'][i] * model.J[i]()) + (AFF * model.REN_SOLAR[i]() * model.CPX_2['SOLAR'][i])) for s in model.S), 3)
    energy_planning.loc['HYDRO', 'Total Cost (mil USD/y)'] = round(sum(((model.REN_HYDRO[i]() * model.REN_COST['HYDRO'][i]) + (AFF * model.CPX_1['HYDRO'][i] * model.K[i]()) + (AFF * model.REN_HYDRO[i]() * model.CPX_2['HYDRO'][i])) for s in model.S), 3)
    energy_planning.loc['BIOMASS', 'Total Cost (mil USD/y)'] = round(sum(((model.REN_BM[i]() * model.REN_COST['BIOMASS'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.L[i]()) + (AFF * model.REN_BM[i]() * model.CPX_2['BIOMASS'][i])) for s in model.S), 3)
    energy_planning.loc['BIOGAS', 'Total Cost (mil USD/y)'] = round(sum(((model.REN_BG[i]() * model.REN_COST['BIOGAS'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.M[i]()) + (AFF * model.REN_BG[i]() * model.CPX_2['BIOGAS'][i])) for s in model.S), 3)
    energy_planning.loc['MSW', 'Total Cost (mil USD/y)'] = round(sum(((model.REN_MSW[i]() * model.REN_COST['MSW'][i]) + (AFF * model.CPX_1['MSW'][i] * model.N[i]()) + (AFF * model.REN_MSW[i]() * model.CPX_2['MSW'][i])) for s in model.S), 3)
    #Below lines are commented by the if/else statement if CT is turned off
    if CT == 'ON':
        energy_planning.loc['TOTAL', 'Emissions bought from previous period (Mt/y)'] = round(sum(model.ET_plant_new [i,s]() for s in model.S), 3)
        energy_planning.loc['TOTAL', 'Emission Rights Bought by Each Plant from government (Mt/y)'] = round(sum(model.CT_emissions [i,s]() for s in model.S), 3)
        energy_planning.loc['TOTAL', 'Emissions rights available for trading within companies (Mt/y)'] = round(sum(model.ET_plant[i,s]() for s in model.S), 3)
        energy_planning.loc['TOTAL', 'Emissions traded within a company (Mt/y)'] = round(sum(model.ET_plant_final [i,s]() for s in model.S), 3)
        energy_planning.loc['TOTAL', 'CT Cost Paid to government to buy emissions from previous period (mil USD / y)'] = round(sum((model.ET_plant_new[i,s]() * model.emissions_unit_cost_new[i,s]())for s in model.S) , 3)
        energy_planning.loc['TOTAL', 'CT Cost Paid by Plant to government (mil USD / y)'] = round(sum((model.CT_emissions [i,s]() * model.emissions_unit_cost_new[i,s]()) for s in model.S), 3)
        energy_planning.loc['TOTAL', 'CT Cost Paid/earned by each plant as a result of trading with each other (mil USD / y)'] = round(sum((model.emissions_unit_cost_new[i,s]() * model.ET_plant_final [i,s]()) for s in model.S), 3)
        energy_planning.loc['TOTAL', 'Balance Emissions to be traded in next period'] = round(model.CT_balance_emissions[i](), 3)
        energy_planning.loc['TOTAL', 'Price paid/earned for balance emissions (mil USD / y)'] = round(sum((model.Emissions_left[i,s]() * model.emissions_unit_cost_new[i,s]()) for s in model.S) , 3)
    else:
        pass
    #Above lines are commented by the if/else statement if CT is turned off
    energy_planning.loc['TOTAL', 'Cumulative total fuel and annualized capital cost (mil USD/y)'] = round(sum(model.energy_cost[i,s]() for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Annualized capital cost (mil USD/y)'] = round(sum(model.annualized_capital_cost[i,s]() for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Fuel cost (mil USD/y)'] = round(sum(model.fuel_cost[i,s]() for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Cost contribution of CCS_1 (mil USD/y)'] = round(sum(((model.net_energy_CCS_1[i,s]() * model.CCS_data['Cost_CCS_1'][i]) + (AFF * model.CCS_data['FX_Cost_CCS_1'][i] * model.B[i,s]())) for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Cost contribution of CCS_2 (mil USD/y)'] = round(sum(((model.net_energy_CCS_2[i,s]() * model.CCS_data['Cost_CCS_2'][i]) + (AFF * model.CCS_data['FX_Cost_CCS_2'][i] * model.C[i,s]())) for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Cost contribution of Solid_1 (mil USD/y)'] = round(sum(((model.solid_1[i,s]() * model.SLD_COST['SOLID_1'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.O[i,s]())) for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Cost contribution of Solid_2 (mil USD/y)'] = round(sum(((model.solid_2[i,s]() * model.SLD_COST['SOLID_2'][i]) + (AFF * model.CPX_1['BIOMASS'][i] * model.P[i,s]())) for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Cost contribution of Gas_1 (mil USD/y)'] = round(sum(((model.gas_1[i,s]() * model.GAS_COST['GAS_1'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.Q[i,s]())) for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Cost contribution of Gas_2 (mil USD/y)'] = round(sum(((model.gas_2[i,s]() * model.GAS_COST['GAS_2'][i]) + (AFF * model.CPX_1['BIOGAS'][i] * model.R[i,s]())) for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Earning by CO2 selling (mil USD / y)'] = round(sum(model.CCS_selling_price[i,s]() for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Earning by selling electricty (mil USD / y)'] = round(sum(model.Plant_earning[i,s]() for s in model.S), 3)
    energy_planning.loc['TOTAL', 'Plant Profit (mil USD/y)'] = round(model.total_profit[i](), 3)
    energy_planning.loc['TOTAL', 'Total Cost (mil USD/y)'] = round(model.sum_cost[i](), 3)
            
    writer = pd.ExcelWriter(file_name, engine = 'openpyxl', mode = 'a', if_sheet_exists = 'new')
    #writer = pd.ExcelWriter(file_name, mode = 'a', if_sheet_exists = 'new')
    energy_planning.to_excel(writer, sheet_name = 'Results_Period_1')
    writer.save()
    
    return 
        

model = multiperiod_energy_planning(model, periods)

emissions_each_period = []
for i in periods:
    emissions_each_period.append(model.new_emission[i].value)

cost_each_period = []
for i in periods:
    cost_each_period.append(model.sum_cost[i].value)
    
profit_each_period = []
for i in periods:
    profit_each_period.append(((sum((model.Plant_earning[i,s].value) for s in model.S)) - (model.sum_cost[i].value)))
    

wb = load_workbook(file_name)
sheet_results = wb['FINAL_RESULTS']

df = pd.DataFrame({'Emissions': emissions_each_period})

for index, row in df.iterrows():
    cell = 'B%d'  % (index + 2)
    sheet_results[cell] = row[0]
    
wb.save(file_name)


df_0 = pd.DataFrame({'Cost': cost_each_period})

for index0, row0 in df_0.iterrows():
    cell0 = 'C%d'  % (index0 + 2)
    sheet_results[cell0] = row0[0]
    
wb.save(file_name)

df_1 = pd.DataFrame({'Profit': profit_each_period})

for index1, row1 in df_1.iterrows():
    cell1 = 'D%d' % (index1 + 2)
    sheet_results[cell1] = row1[0]

wb.save(file_name)

#Below lines print/display the results in excel-based user interface
if Print_Result == 'YES':
    for i in list(range(1, numperiods,1)):
        multiperiod_energy_planning_results(model,i)
else:
    pass
